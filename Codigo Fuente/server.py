# ===========================================================
#  SmartCar - Servidor Principal con Seguridad JWT
# ===========================================================
import sys
import os
from flask import Flask, jsonify, request, render_template, send_from_directory, send_file
from flask_cors import CORS
from datetime import datetime, timedelta
import jwt
from functools import wraps
from io import BytesIO
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# Configurar rutas del sistema
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), 'backend')))

# ===========================================================
# IMPORTACIONES DEL BACKEND
# ===========================================================
from backend.core.db.connection import get_connection
from backend.models.user_model import verificar_usuario
from backend.core.auditoria_utils import registrar_auditoria_global # Nueva Auditoría Centralizada

# Controladores
from backend.core.controller_personas import (
    desactivar_persona_controller,
    obtener_personas_controller,
    crear_persona_controller,
    actualizar_persona_controller
)
from backend.core.controller_vehiculos import (
    eliminar_vehiculo_controller,
    obtener_vehiculos_controller,
    crear_vehiculo_controller,
    actualizar_vehiculo_controller
)
from backend.core.controller_accesos import (
    obtener_historial_accesos, 
    procesar_validacion_acceso
)
from backend.core.controller_calendario import (
    obtener_eventos_controller,
    crear_evento_controller,
    actualizar_evento_controller,
    eliminar_evento_controller,
    verificar_evento_controller
)
from backend.core.controller_incidencias import (
    obtener_vehiculos_en_patio, 
    crear_incidente_manual,
    obtener_estado_actual_patio, # Nuevo
    crear_novedad_general,       # Nuevo
    obtener_historial_vigilante  # Nuevo
)
from backend.core.controller_alertas import (
    obtener_alertas_controller, 
    eliminar_alerta_controller
)
from backend.models.auditoria import obtener_historial_auditoria
from backend.models.dashboard_model import (
    obtener_ultimos_accesos,
    contar_total_vehiculos,
    contar_alertas_activas,
    buscar_placa_bd
)
from backend.models.admin_model import (
    obtener_datos_dashboard,
    obtener_accesos_detalle,
    registrar_vigilante,
    obtener_data_reporte_completo # <--- NUEVA IMPORTACIÓN
)
# ===========================================================
# CONFIGURACIÓN FLASK
# ===========================================================
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "frontend", "templates")
STATIC_DIR = os.path.join(BASE_DIR, "frontend", "static")

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
CORS(app)

app.config["SECRET_KEY"] = "SmartCar_SeguridadUltra_2025"

# ===========================================================
# DECORADOR: VALIDAR TOKEN JWT
# ===========================================================
def token_requerido(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({"error": "Token no proporcionado"}), 401

        try:
            token = token.replace("Bearer ", "")
            datos = jwt.decode(token, app.config["SECRET_KEY"], algorithms=["HS256"])
            request.usuario_actual = datos
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token expirado"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "Token inválido"}), 401

        return f(*args, **kwargs)
    return decorador

# ===========================================================
# RUTAS PÚBLICAS & LOGIN
# ===========================================================
@app.route("/")
def index():
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def login():
    try:
        data = request.get_json()
        usuario = data.get("usuario")
        clave = data.get("clave")
        rol = data.get("rol")

        if not usuario or not clave or not rol:
            return jsonify({"error": "Faltan campos requeridos"}), 400

        user = verificar_usuario(usuario, clave, rol)
        if not user:
            return jsonify({"error": "Usuario, clave o rol incorrectos"}), 401

        # Generar Token
        token = jwt.encode({
            "usuario": user["usuario"],
            "rol": user["rol"],
            "id_audit": user["id_audit"], # ID real del usuario (nu)
            "exp": datetime.utcnow() + timedelta(hours=8)
        }, app.config["SECRET_KEY"], algorithm="HS256")
        
        # --- AUDITORÍA LOGIN ---
        registrar_auditoria_global(
            id_usuario=user["id_audit"],
            entidad="SISTEMA",
            id_entidad=0,
            accion="INICIO_SESION",
            datos_nuevos={"usuario": user["usuario"], "rol": user["rol"]}
        )

        return jsonify({
            "status": "ok",
            "token": token,
            "user": {
                "nombre": user["nombre"],
                "usuario": user["usuario"],
                "rol": user["rol"]
            }
        }), 200

    except Exception as e:
        print("❌ Error en login:", e)
        return jsonify({"error": "Error interno del servidor"}), 500

# ===========================================================
# RUTAS DASHBOARD VIGILANTE (Info Básica)
# ===========================================================
@app.route("/api/ultimos_accesos", methods=["GET"])
def api_ultimos_accesos():
    return jsonify(obtener_ultimos_accesos())

@app.route("/api/total_vehiculos", methods=["GET"])
def api_total_vehiculos():
    return jsonify(contar_total_vehiculos())

@app.route("/api/alertas_activas", methods=["GET"])
def api_alertas_activas():
    return jsonify(contar_alertas_activas())

@app.route("/api/buscar_placa/<placa>", methods=["GET"])
def api_buscar_placa(placa):
    data = buscar_placa_bd(placa)
    if data:
        return jsonify(data)
    else:
        return jsonify({"error": "Placa no encontrada"}), 404

# ===========================================================
# RUTAS DASHBOARD ADMIN
# ===========================================================
@app.route("/api/admin/resumen", methods=["GET"])
@token_requerido
def api_admin_resumen():
    return jsonify(obtener_datos_dashboard())

@app.route("/api/admin/accesos", methods=["GET"])
@token_requerido
def api_admin_accesos():
    return jsonify(obtener_accesos_detalle())

@app.route("/api/admin/registrar_vigilante", methods=["POST"])
@token_requerido
def api_registrar_vigilante():
    data = request.get_json()
    ok = registrar_vigilante(
        data.get("nombre"),
        data.get("doc_identidad"),
        data.get("telefono"),
        data.get("id_rol")
    )
    if ok:
        return jsonify({"status": "ok"})
    return jsonify({"error": "No se pudo registrar"}), 500

@app.route("/api/admin/auditoria", methods=["GET"])
@token_requerido
def api_admin_auditoria():
    if request.usuario_actual.get('rol') != 'Administrador':
        return jsonify({"error": "Acceso no autorizado"}), 403
    try:
        historial = obtener_historial_auditoria()
        return jsonify(historial), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ===========================================================
# REPORTES (PDF / EXCEL)
# ===========================================================
@app.route("/api/admin/exportar/excel", methods=["GET"])
@token_requerido
def exportar_excel():
    try:
        # 1. Obtener fechas del filtro
        fecha_inicio = request.args.get('inicio', datetime.now().strftime('%Y-%m-%d'))
        fecha_fin = request.args.get('fin', datetime.now().strftime('%Y-%m-%d'))

        # 2. Obtener Data Completa
        reporte = obtener_data_reporte_completo(fecha_inicio, fecha_fin)
        if not reporte:
            return jsonify({"error": "Error obteniendo datos"}), 500

        wb = Workbook()
        
        # --- HOJA 1: RESUMEN Y ESTADÍSTICAS ---
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Gerencial"
        ws_resumen.append(["REPORTE OPERATIVO SMARTCAR", f"Del {fecha_inicio} al {fecha_fin}"])
        ws_resumen.append([])
        
        stats = reporte["estadisticas"]
        pico = reporte["hora_pico"]
        ws_resumen.append(["METRICA", "VALOR"])
        ws_resumen.append(["Total Movimientos", stats['total_movimientos']])
        ws_resumen.append(["Accesos Autorizados", stats['autorizados']])
        ws_resumen.append(["Accesos Denegados", stats['denegados']])
        if pico:
            ws_resumen.append(["Hora Pico", f"{int(pico['hora'])}:00 - {int(pico['hora'])+1}:00 ({pico['cantidad']} vehs)"])

        # --- HOJA 2: NOVEDADES VIGILANTES ---
        ws_nov = wb.create_sheet("Novedades Vigilancia")
        ws_nov.append(["Fecha", "Vigilante", "Asunto", "Descripción"])
        for n in reporte["novedades"]:
            ws_nov.append([n['fecha'], n['vigilante'], n['asunto'], n['descripcion']])

        # --- HOJA 3: INCIDENTES RESUELTOS ---
        ws_inc = wb.create_sheet("Incidentes Resueltos")
        ws_inc.append(["Fecha Resolución", "Resolutor", "Tipo Alerta", "Acción Tomada"])
        import json
        for a in reporte["alertas_resueltas"]:
            # Parseamos los JSON strings guardados en auditoria
            prev = json.loads(a['datos_previos']) if isinstance(a['datos_previos'], str) else a['datos_previos']
            new = json.loads(a['datos_nuevos']) if isinstance(a['datos_nuevos'], str) else a['datos_nuevos']
            
            tipo_alerta = prev.get('tipo', 'Desconocido') if prev else 'N/A'
            resolucion = new.get('resolucion', 'Sin detalle') if new else 'N/A'
            
            ws_inc.append([a['fecha_resolucion'], a['resolutor'], tipo_alerta, resolucion])

        # --- HOJA 4: DETALLE ACCESOS ---
        ws_acc = wb.create_sheet("Log Accesos")
        ws_acc.append(["Fecha", "Placa", "Tipo", "Resultado", "Vigilante"])
        for acc in reporte["accesos"]:
            ws_acc.append([acc['fecha'], acc['placa'], acc['tipo'], acc['resultado'], acc['vigilante']])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"Reporte_Completo_{fecha_inicio}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/exportar/pdf", methods=["GET"])
@token_requerido
def exportar_pdf():
    try:
        fecha_inicio = request.args.get('inicio', datetime.now().strftime('%Y-%m-%d'))
        fecha_fin = request.args.get('fin', datetime.now().strftime('%Y-%m-%d'))
        
        reporte = obtener_data_reporte_completo(fecha_inicio, fecha_fin)
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # --- HEADER ---
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "SmartCar - Informe Gerencial Operativo")
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 70, f"Periodo: {fecha_inicio} al {fecha_fin}")

        # --- ESTADÍSTICAS ---
        y = height - 120
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "1. Resumen Estadístico")
        y -= 20
        p.setFont("Helvetica", 10)
        stats = reporte["estadisticas"]
        p.drawString(70, y, f"• Total Movimientos: {stats['total_movimientos']}")
        y -= 15
        p.drawString(70, y, f"• Accesos Autorizados: {stats['autorizados']}")
        y -= 15
        p.drawString(70, y, f"• Accesos Denegados: {stats['denegados']}")
        y -= 15
        if reporte["hora_pico"]:
            hp = reporte["hora_pico"]
            p.drawString(70, y, f"• Hora Pico: {int(hp['hora'])}:00 horas ({hp['cantidad']} accesos)")
        
        # --- NOVEDADES ---
        y -= 40
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "2. Novedades y Observaciones (Últimas 10)")
        y -= 20
        p.setFont("Helvetica", 9)
        for n in reporte["novedades"][:10]: # Limitamos a 10 para que quepa en 1 hoja demo
            texto = f"{n['fecha']} - {n['vigilante']}: {n['asunto']} - {n['descripcion'][:50]}..."
            p.drawString(70, y, texto)
            y -= 15
            if y < 50: 
                p.showPage()
                y = height - 50

        # --- ALERTAS RESUELTAS ---
        y -= 20
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "3. Resolución de Incidentes")
        y -= 20
        p.setFont("Helvetica", 9)
        import json
        for a in reporte["alertas_resueltas"][:10]:
            prev = json.loads(a['datos_previos']) if isinstance(a['datos_previos'], str) else a['datos_previos']
            new = json.loads(a['datos_nuevos']) if isinstance(a['datos_nuevos'], str) else a['datos_nuevos']
            tipo = prev.get('tipo', 'N/A') if prev else 'N/A'
            resol = new.get('resolucion', 'N/A') if new else 'N/A'
            
            texto = f"{a['fecha_resolucion']} - {tipo} -> Acción: {resol} (Por: {a['resolutor']})"
            p.drawString(70, y, texto)
            y -= 15

        p.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="Informe_Gerencial.pdf", mimetype="application/pdf")

    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500

# ===========================================================
# CRUD PERSONAS
# ===========================================================
@app.route("/api/personas", methods=["GET"])
@token_requerido
def get_personas():
    try:
        personas = obtener_personas_controller()
        return jsonify(personas), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/personas", methods=["POST"])
@token_requerido
def create_persona():
    try:
        data = request.json
        nuevo_id = crear_persona_controller(data, request.usuario_actual)
        return jsonify({"mensaje": "Persona creada", "id_persona": nuevo_id}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/personas/<int:id_persona>", methods=["PUT"])
@token_requerido
def update_persona(id_persona):
    try:
        data = request.json
        actualizar_persona_controller(id_persona, data, request.usuario_actual)
        return jsonify({"mensaje": "Persona actualizada"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route("/api/personas/<int:id_persona>", methods=["DELETE"])
@token_requerido
def delete_persona(id_persona):
    try:
        desactivar_persona_controller(id_persona, request.usuario_actual)
        return jsonify({"mensaje": "Persona desactivada"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ===========================================================
# CRUD VEHÍCULOS
# ===========================================================
@app.route("/api/vehiculos", methods=["GET"])
@token_requerido
def get_vehiculos():
    try:
        vehiculos = obtener_vehiculos_controller()
        return jsonify(vehiculos), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vehiculos", methods=["POST"])
@token_requerido
def create_vehiculo():
    try:
        data = request.json
        nuevo_id = crear_vehiculo_controller(data, request.usuario_actual)
        return jsonify({"mensaje": "Vehículo creado", "id_vehiculo": nuevo_id}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vehiculos/<int:id_vehiculo>", methods=["PUT"])
@token_requerido
def update_vehiculo(id_vehiculo):
    try:
        data = request.json
        actualizar_vehiculo_controller(id_vehiculo, data, request.usuario_actual)
        return jsonify({"mensaje": "Vehículo actualizado"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route("/api/vehiculos/<int:id_vehiculo>", methods=["DELETE"])
@token_requerido
def delete_vehiculo(id_vehiculo):
    try:
        eliminar_vehiculo_controller(id_vehiculo, request.usuario_actual)
        return jsonify({"mensaje": "Vehículo eliminado"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ================================================
@app.route("/api/usuario", methods=["GET"])
@token_requerido
def obtener_usuario():
    datos = request.usuario_actual
    return jsonify({"status": "ok", "user": datos}), 200

# ===========================================================
# CONTROL DE ACCESOS (HISTORIAL Y OCR)
# ===========================================================
@app.route("/api/accesos", methods=["GET"])
@token_requerido
def get_historial_accesos():
    try:
        filtros = {
            "placa": request.args.get('placa'),
            "tipo": request.args.get('tipo'),
            "desde": request.args.get('desde'),
            "hasta": request.args.get('hasta')
        }
        historial = obtener_historial_accesos(filtros)
        return jsonify(historial), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/accesos/validar", methods=["POST"])
def validar_acceso_ocr():
    try:
        respuesta, status = procesar_validacion_acceso(request.data, vigilante_id=1)
        return jsonify(respuesta), status
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ===========================================================
# ALERTAS (ADMIN)
# ===========================================================
@app.route("/api/admin/alertas", methods=["GET"])
@token_requerido
def get_alertas():
    try:
        alertas = obtener_alertas_controller()
        return jsonify(alertas), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/alertas/<int:id_alerta>", methods=["DELETE"])
@token_requerido
def delete_alerta(id_alerta):
    try:
        # Obtenemos la acción de resolución enviada desde el frontend (JSON body)
        # Si no envían nada, usamos "Resolución General"
        data = request.get_json() or {}
        accion = data.get('accion_resolucion', 'Resolución General')

        if eliminar_alerta_controller(id_alerta, request.usuario_actual, accion):
            return jsonify({"mensaje": "Alerta resuelta y archivada"}), 200
        else:
            return jsonify({"error": "No se pudo resolver la alerta"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ===========================================================
# NOVEDADES Y REPORTES (VIGILANTE)
# ===========================================================
@app.route("/api/vigilante/vehiculos-en-patio", methods=["GET"])
@token_requerido
def get_vehiculos_patio():
    try:
        vehiculos = obtener_vehiculos_en_patio()
        return jsonify(vehiculos), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vigilante/reportar", methods=["POST"])
@token_requerido
def post_reportar_incidente():
    try:
        data = request.get_json()
        id_vigilante = request.usuario_actual['id_audit']
        if crear_incidente_manual(data, id_vigilante):
            return jsonify({"mensaje": "Reportado"}), 201
        return jsonify({"error": "Error"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/vigilante/estado-patio", methods=["GET"])
@token_requerido
def get_estado_patio():
    """Devuelve contador de vehículos y hora del servidor"""
    data = obtener_estado_actual_patio()
    return jsonify(data), 200

@app.route("/api/vigilante/novedad", methods=["POST"])
@token_requerido
def post_novedad():
    """Crea una novedad general (ej: Daño infraestructura)"""
    data = request.get_json()
    id_vigilante = request.usuario_actual['id_audit']
    if crear_novedad_general(data, id_vigilante):
        return jsonify({"mensaje": "Novedad registrada"}), 201
    return jsonify({"error": "Error al registrar"}), 500

@app.route("/api/vigilante/mis-reportes", methods=["GET"])
@token_requerido
def get_mis_reportes():
    """Obtiene el historial de reportes del vigilante logueado"""
    id_vigilante = request.usuario_actual['id_audit']
    data = obtener_historial_vigilante(id_vigilante)
    return jsonify(data), 200

# ===========================================================
# CALENDARIO Y EVENTOS
# ===========================================================
@app.route("/api/eventos", methods=["GET"])
@token_requerido
def get_eventos():
    try:
        eventos = obtener_eventos_controller()
        return jsonify(eventos), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/eventos", methods=["POST"])
@token_requerido
def create_evento():
    if request.usuario_actual.get('rol') != 'Administrador':
        return jsonify({"error": "No autorizado"}), 403
    try:
        data = request.get_json()
        id_nuevo = crear_evento_controller(data, request.usuario_actual)
        return jsonify({"mensaje": "Evento creado", "id": id_nuevo}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/eventos/<int:id_evento>", methods=["PUT"])
@token_requerido
def update_evento(id_evento):
    if request.usuario_actual.get('rol') != 'Administrador':
        return jsonify({"error": "No autorizado"}), 403
    try:
        data = request.get_json()
        actualizar_evento_controller(id_evento, data)
        return jsonify({"mensaje": "Evento actualizado"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/eventos/<int:id_evento>", methods=["DELETE"])
@token_requerido
def delete_evento(id_evento):
    if request.usuario_actual.get('rol') != 'Administrador':
        return jsonify({"error": "No autorizado"}), 403
    try:
        # Pasamos usuario_actual para la auditoría
        if eliminar_evento_controller(id_evento, request.usuario_actual):
            return jsonify({"mensaje": "Evento eliminado correctamente"}), 200
        else:
            return jsonify({"error": "Evento no encontrado o no se pudo eliminar"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/eventos/<int:id_evento>/verificar", methods=["PUT"])
@token_requerido
def verify_evento(id_evento):
    try:
        data = request.get_json()
        verificar_evento_controller(id_evento, data.get('verificado', True))
        return jsonify({"mensaje": "Verificado"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ===========================================================
# RUTA PARA ARCHIVOS ESTÁTICOS (Frontend)
# ===========================================================
@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory(app.static_folder, filename)

# ===========================================================
# INICIO DEL SERVIDOR
# ===========================================================
if __name__ == "__main__":
    print("✅ Servidor SmartCar ejecutándose en http://127.0.0.1:5000")
    app.run(host="127.0.0.1", port=5000, debug=True)